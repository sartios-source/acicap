from typing import Dict, Any, List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill


def _apply_header(ws, row: int, headers: List[str]) -> None:
    header_fill = PatternFill("solid", fgColor="1E88E5")
    for idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=idx, value=title)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")


def export_fabric_excel(fabric_name: str, analysis: Dict[str, Any]) -> Workbook:
    wb = Workbook()

    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Fabric: {fabric_name}"
    ws["A1"].font = Font(size=14, bold=True)

    summary = analysis.get("summary", {})
    port = analysis.get("ports", {})
    l3out = analysis.get("l3out", {})
    vpc = analysis.get("vpc", {})
    vlan_pools = analysis.get("vlan_pools", {})

    rows = [
        ("Leafs", summary.get("leafs", 0)),
        ("Spines", summary.get("spines", 0)),
        ("FEX", summary.get("fex", 0)),
        ("Tenants", summary.get("tenants", 0)),
        ("VRFs", summary.get("vrfs", 0)),
        ("BDs", summary.get("bds", 0)),
        ("EPGs", summary.get("epgs", 0)),
        ("Subnets", summary.get("subnets", 0)),
        ("Contracts", summary.get("contracts", 0)),
        ("Total Ports", port.get("total", 0)),
        ("Ports Up", port.get("up", 0)),
        ("Ports Down", port.get("down", 0)),
        ("Ports Unknown", port.get("unknown", 0)),
        ("Ports With EPG", port.get("ports_with_epg", 0)),
        ("L3Outs", l3out.get("l3outs", 0)),
        ("External EPGs", l3out.get("external_epgs", 0)),
        ("Border Leafs", l3out.get("border_leaf_count", 0)),
        ("BGP Peers", l3out.get("bgp_peers", 0)),
        ("OSPF Interfaces", l3out.get("ospf_interfaces", 0)),
        ("vPC Domains", vpc.get("vpc_domains", 0)),
        ("Port-Channels", vpc.get("port_channels", 0)),
        ("VLAN Pools", vlan_pools.get("pool_count", 0)),
        ("VLAN Pool Capacity", vlan_pools.get("pool_vlan_capacity", 0)),
        ("VLANs Used", vlan_pools.get("used_vlan_count", 0)),
    ]
    _apply_header(ws, 2, ["Metric", "Value"])
    for metric, value in rows:
        ws.append([metric, value])

    tenant_ws = wb.create_sheet("Tenants")
    tenant_rows = analysis.get("tenants", {}).get("rows", [])
    _apply_header(tenant_ws, 1, ["Tenant", "VRFs", "BDs", "EPGs", "Subnets", "Contracts"])
    for row in tenant_rows:
        tenant_ws.append([
            row.get("tenant", ""),
            row.get("vrfs", 0),
            row.get("bds", 0),
            row.get("epgs", 0),
            row.get("subnets", 0),
            row.get("contracts", 0),
        ])

    epg_ws = wb.create_sheet("EPG Spread")
    epg_rows = analysis.get("epg_spread", {}).get("rows", [])
    _apply_header(epg_ws, 1, ["Tenant", "EPG", "Path Count", "Node Count"])
    for row in epg_rows:
        epg_ws.append([
            row.get("tenant", ""),
            row.get("epg", ""),
            row.get("path_count", 0),
            row.get("node_count", 0),
        ])

    vlan_ws = wb.create_sheet("VLAN Overlap")
    overlaps = analysis.get("vlan_overlap", {}).get("overlaps", [])
    _apply_header(vlan_ws, 1, ["VLAN", "Tenant Count", "Tenants"])
    for row in overlaps:
        vlan_ws.append([
            row.get("vlan", ""),
            row.get("tenant_count", 0),
            ", ".join(row.get("tenants", []))
        ])

    return wb
